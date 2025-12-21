"""
Main orchestrator for JSON conversion tool.
CLI interface with progress tracking and comprehensive error reporting.
"""

import argparse
import sys
import shutil
from pathlib import Path
from datetime import datetime
from typing import List
import json

try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False
    print("Warning: tqdm not installed. Progress bar will not be available.")
    print("Install with: pip install tqdm")

try:
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not installed. Excel reports will not be available.")
    print("Install with: pip install openpyxl")

from SCRIPTS.config import EXCEL_COLUMNS, ERROR_TYPES
from SCRIPTS.utils import load_json_file, save_json_file, ValidationError, ConversionError, format_timestamp, detect_question_types
from SCRIPTS.pre_validator import validate_pre_conversion
from SCRIPTS.converter import convert_question
from SCRIPTS.post_validator import validate_post_conversion


class ConversionStats:
    """Track conversion statistics"""
    def __init__(self):
        self.total = 0
        self.success = 0
        self.pre_validation_failed = 0
        self.conversion_failed = 0
        self.post_validation_failed = 0
        self.errors = []  # List of error dictionaries for Excel report
        self.warnings = []  # List of warning dictionaries for Excel report
        
    def add_error(self, filename: str, question_id: str, error_type: str, 
                  error_message: str, field: str = "", actual_value: str = "", 
                  expected: str = ""):
        """Add an error to the error log"""
        self.errors.append({
            "filename": filename,
            "question_id": question_id,
            "error_type": error_type,
            "error_message": error_message,
            "field": field,
            "actual_value": str(actual_value),
            "expected": expected,
            "timestamp": format_timestamp()
        })
    
    def add_warning(self, filename: str, question_id: str, warning_message: str):
        """Add a warning to the warning log"""
        self.warnings.append({
            "filename": filename,
            "question_id": question_id,
            "warning_message": warning_message,
            "timestamp": format_timestamp()
        })


def discover_json_files(input_path: Path, filter_types: List[str] = None) -> List[Path]:
    """
    Discover all JSON files in input path.
    
    Args:
        input_path: Root directory to search
        filter_types: Optional list of question types to filter by
        
    Returns:
        List of JSON file paths
    """
    json_files = []
    
    if input_path.is_file():
        if input_path.suffix == '.json':
            json_files.append(input_path)
    else:
        json_files = list(input_path.rglob('*.json'))
    
    # Filter by type if requested
    if filter_types:
        filtered = []
        for filepath in json_files:
            try:
                data = load_json_file(filepath)
                types = detect_question_types(data)
                # Include if any part type matches filter
                if any(t in filter_types for t in types):
                    filtered.append(filepath)
            except:
                pass  # Skip files that can't be loaded
        json_files = filtered
    
    return json_files


def process_file(filepath: Path, output_dir: Path, pre_validation_failed_dir: Path, 
                 failed_dir: Path, post_validation_failed_dir: Path, stats: ConversionStats, dry_run: bool = False, 
                 verbose: bool = False) -> bool:
    """
    Process a single JSON file through the conversion pipeline.
    
    Args:
        filepath: Path to the JSON file
        output_dir: Directory for converted files
        pre_validation_failed_dir: Directory for pre-validation failed files
        failed_dir: Directory for failed conversions (conversion errors)
        post_validation_failed_dir: Directory for post-validation failed files
        stats: Statistics tracker
        dry_run: If True, don't write files
        verbose: If True, print detailed messages
        
    Returns:
        True if successful, False otherwise
    """
    filename = filepath.name
    stats.total += 1
    
    try:
        # Load JSON file
        json_data = load_json_file(filepath)
        question_id = str(json_data.get('question_id', 'unknown'))
        
        # Pre-conversion validation
        is_valid, errors, warnings = validate_pre_conversion(json_data, filename)
        
        # Log warnings (even if validation passes)
        for warning in warnings:
            stats.add_warning(filename, question_id, warning)
        
        if not is_valid:
            stats.pre_validation_failed += 1
            
            # Copy pre-validation failed file to pre_validation_failed/ folder (unchanged)
            if not dry_run:
                pre_validation_failed_path = pre_validation_failed_dir / filename
                shutil.copy2(filepath, pre_validation_failed_path)
            
            # Log all errors
            for error in errors:
                stats.add_error(
                    filename, question_id,
                    ERROR_TYPES['PRE_VALIDATION'],
                    error
                )
            
            if verbose:
                print(f"  PRE-VALIDATION FAILED: {filename} - {len(errors)} validation errors")
            
            return False
        
        # Convert
        try:
            converted_json = convert_question(json_data, filename)
        except (ValidationError, ConversionError) as e:
            stats.conversion_failed += 1
            
            # Copy failed file to failed/ folder (unchanged)
            if not dry_run:
                failed_path = failed_dir / filename
                shutil.copy2(filepath, failed_path)
            
            error_msg = str(e)
            field = getattr(e, 'field', '')
            actual = getattr(e, 'actual_value', '')
            expected = getattr(e, 'expected', '')
            
            stats.add_error(
                filename, question_id,
                ERROR_TYPES['CONVERSION'],
                error_msg, field, actual, expected
            )
            
            if verbose:
                print(f"  CONVERSION FAILED: {filename} - {error_msg}")
            
            return False
        
        # Post-conversion validation
        is_valid_post, post_errors = validate_post_conversion(converted_json)
        
        if not is_valid_post:
            stats.post_validation_failed += 1

            # Save post-validation failed file to post_validation_failed/ folder
            if not dry_run:
                post_validation_failed_path = post_validation_failed_dir / filename
                save_json_file(converted_json, post_validation_failed_path)
            
            # Log all post-validation errors
            for error in post_errors:
                stats.add_error(
                    filename, question_id,
                    ERROR_TYPES['POST_VALIDATION'],
                    error
                )
            
            if verbose:
                print(f"  POST-VALIDATION FAILED: {filename} - {len(post_errors)} errors")
            
            return False
        
        # Success - save converted file
        if not dry_run:
            output_path = output_dir / filename
            save_json_file(converted_json, output_path)
        
        stats.success += 1
        
        if verbose:
            types = detect_question_types(json_data)
            print(f"  SUCCESS: {filename} (types: {', '.join(types)})")
        
        return True
        
    except Exception as e:
        stats.conversion_failed += 1
        
        try:
            question_id = str(json_data.get('question_id', 'unknown'))
        except:
            question_id = 'unknown'

        # Copy failed file to failed/ folder (unchanged)
        if not dry_run:
            failed_path = failed_dir / filename
            shutil.copy2(filepath, failed_path)

        stats.add_error(
            filename, question_id,
            ERROR_TYPES['CONVERSION'],
            f"Unexpected error: {str(e)}"
        )
        
        if verbose:
            print(f"  ERROR: {filename} - {str(e)}")
        
        return False


def generate_excel_report(stats: ConversionStats, output_path: Path):
    """
    Generate Excel report of all errors and warnings.
    
    Args:
        stats: Statistics with error and warning lists
        output_path: Path to save the Excel file
    """
    if not HAS_OPENPYXL:
        print("Cannot generate Excel report: openpyxl not installed")
        return
    
    wb = Workbook()
    
    # Errors sheet
    ws_errors = wb.active
    ws_errors.title = "Errors"
    
    # Write error headers
    headers = EXCEL_COLUMNS
    for col_idx, header in enumerate(headers, 1):
        ws_errors.cell(row=1, column=col_idx, value=header)
    
    # Write errors
    for row_idx, error in enumerate(stats.errors, 2):
        ws_errors.cell(row=row_idx, column=1, value=error['filename'])
        ws_errors.cell(row=row_idx, column=2, value=error['question_id'])
        ws_errors.cell(row=row_idx, column=3, value=error['error_type'])
        ws_errors.cell(row=row_idx, column=4, value=error['error_message'])
        ws_errors.cell(row=row_idx, column=5, value=error['field'])
        ws_errors.cell(row=row_idx, column=6, value=error['actual_value'])
        ws_errors.cell(row=row_idx, column=7, value=error['expected'])
        ws_errors.cell(row=row_idx, column=8, value=error['timestamp'])
    
    # Auto-adjust column widths for errors
    for column in ws_errors.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_errors.column_dimensions[column_letter].width = adjusted_width
    
    # Warnings sheet (if there are warnings)
    if len(stats.warnings) > 0:
        ws_warnings = wb.create_sheet(title="Warnings")
        
        # Write warning headers
        warning_headers = ["Filename", "Question ID", "Warning Message", "Timestamp"]
        for col_idx, header in enumerate(warning_headers, 1):
            ws_warnings.cell(row=1, column=col_idx, value=header)
        
        # Write warnings
        for row_idx, warning in enumerate(stats.warnings, 2):
            ws_warnings.cell(row=row_idx, column=1, value=warning['filename'])
            ws_warnings.cell(row=row_idx, column=2, value=warning['question_id'])
            ws_warnings.cell(row=row_idx, column=3, value=warning['warning_message'])
            ws_warnings.cell(row=row_idx, column=4, value=warning['timestamp'])
        
        # Auto-adjust column widths for warnings
        for column in ws_warnings.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_warnings.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)
    print(f"\nExcel report saved to: {output_path}")
    if len(stats.warnings) > 0:
        print(f"  - {len(stats.errors)} errors in 'Errors' sheet")
        print(f"  - {len(stats.warnings)} warnings in 'Warnings' sheet")


def generate_text_log(stats: ConversionStats, output_path: Path):
    """
    Generate text log of conversion process.
    
    Args:
        stats: Statistics object
        output_path: Path to save the log file
    """
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write("JSON CONVERSION LOG\n")
        f.write("=" * 80 + "\n\n")
        f.write(f"Generated: {format_timestamp()}\n\n")
        
        f.write("SUMMARY\n")
        f.write("-" * 80 + "\n")
        f.write(f"Total files processed: {stats.total}\n")
        f.write(f"Successfully converted: {stats.success}\n")
        f.write(f"Pre-validation failed: {stats.pre_validation_failed}\n")
        f.write(f"Conversion failed: {stats.conversion_failed}\n")
        f.write(f"Post-validation failed: {stats.post_validation_failed}\n")
        f.write(f"Total failed: {stats.total - stats.success}\n")
        f.write(f"Total warnings: {len(stats.warnings)}\n")
        
        if stats.total > 0:
            success_rate = (stats.success / stats.total) * 100
            f.write(f"Success rate: {success_rate:.2f}%\n")
        
        f.write("\n" + "=" * 80 + "\n")
        f.write("ERRORS\n")
        f.write("=" * 80 + "\n\n")
        
        for error in stats.errors:
            f.write(f"File: {error['filename']}\n")
            f.write(f"Question ID: {error['question_id']}\n")
            f.write(f"Error Type: {error['error_type']}\n")
            f.write(f"Message: {error['error_message']}\n")
            if error['field']:
                f.write(f"Field: {error['field']}\n")
            if error['actual_value']:
                f.write(f"Actual: {error['actual_value']}\n")
            if error['expected']:
                f.write(f"Expected: {error['expected']}\n")
            f.write(f"Timestamp: {error['timestamp']}\n")
            f.write("-" * 80 + "\n\n")
        
        if len(stats.warnings) > 0:
            f.write("\n" + "=" * 80 + "\n")
            f.write("WARNINGS\n")
            f.write("=" * 80 + "\n\n")
            
            for warning in stats.warnings:
                f.write(f"File: {warning['filename']}\n")
                f.write(f"Question ID: {warning['question_id']}\n")
                f.write(f"Warning: {warning['warning_message']}\n")
                f.write(f"Timestamp: {warning['timestamp']}\n")
                f.write("-" * 80 + "\n\n")
    
    print(f"Text log saved to: {output_path}")


def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(
        description='Convert question JSON files from old to new structure',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Convert all files
  python main.py --input ../old_structure_samples/ --output ../converted/
  
  # Convert only specific types
  python main.py --input ../old_structure_samples/ --output ../converted/ --types counting,mcq
  
  # Dry run (no files written)
  python main.py --input ../old_structure_samples/ --output ../converted/ --dry-run
  
  # Verbose output
  python main.py --input ../old_structure_samples/ --output ../converted/ --verbose
        """
    )
    
    parser.add_argument('--input', '-i', type=str,
                       help='Input directory or file path')
    parser.add_argument('--output', '-o', type=str,
                       help='Output directory for converted files')
    parser.add_argument('--types', '-t', type=str,
                       help='Comma-separated list of question types to convert (e.g., counting,mcq)')
    parser.add_argument('--dry-run', action='store_true',
                       help='Run without writing files (validation only)')
    parser.add_argument('--verbose', '-v', action='store_true',
                       help='Print detailed progress messages')
    
    args = parser.parse_args()
    
    # Setup paths
    input_path = Path(args.input).resolve() if args.input else Path('INPUT').resolve()
    output_dir = Path(args.output).resolve() / 'CONVERTED' if args.output else Path('OUTPUTS/CONVERTED').resolve()
    pre_validation_failed_dir = output_dir.parent / 'PRE_CONVERSION_VALIDATION_FAILED'
    failed_dir = output_dir.parent / 'CONVERSION_FAILED'
    post_validation_failed_dir = output_dir.parent / 'POST_CONVERSION_VALIDATION_FAILED'    
    reports_dir = output_dir.parent / 'LOGS_REPORTS'
    logs_dir = output_dir.parent / 'LOGS_REPORTS'
    
    # Create directories
    if not args.dry_run:
        input_path.mkdir(parents=True, exist_ok=True)
        output_dir.mkdir(parents=True, exist_ok=True)
        pre_validation_failed_dir.mkdir(parents=True, exist_ok=True)
        failed_dir.mkdir(parents=True, exist_ok=True)
        post_validation_failed_dir.mkdir(parents=True, exist_ok=True)
        reports_dir.mkdir(parents=True, exist_ok=True)
        logs_dir.mkdir(parents=True, exist_ok=True)
    
    # Validate input
    if not input_path.exists():
        print(f"Error: Input path does not exist: {input_path}")
        sys.exit(1)
    
    # Parse filter types
    filter_types = None
    if args.types:
        filter_types = [t.strip() for t in args.types.split(',')]
        print(f"Filtering by question types: {', '.join(filter_types)}")
    
    # Discover files
    print(f"Discovering JSON files in: {input_path}")
    json_files = discover_json_files(input_path, filter_types)
    print(f"Found {len(json_files)} files to process")
    
    if len(json_files) == 0:
        print("No files to process. Exiting.")
        sys.exit(0)
    
    if args.dry_run:
        print("\n*** DRY RUN MODE - No files will be written ***\n")
    
    # Process files
    stats = ConversionStats()
    
    print("\nProcessing files...")
    
    if HAS_TQDM and not args.verbose:
        # Use progress bar
        for filepath in tqdm(json_files, desc="Converting", unit="file"):
            process_file(filepath, output_dir, pre_validation_failed_dir, failed_dir, post_validation_failed_dir, stats, args.dry_run, args.verbose)
    else:
        # Simple counter
        for i, filepath in enumerate(json_files, 1):
            if not args.verbose:
                if i % 100 == 0 or i == len(json_files):
                    print(f"  Progress: {i}/{len(json_files)} ({i*100//len(json_files)}%)")
            else:
                print(f"[{i}/{len(json_files)}] Processing: {filepath.name}")
            
            process_file(filepath, output_dir, pre_validation_failed_dir, failed_dir, post_validation_failed_dir, stats, args.dry_run, args.verbose)
    
    # Print summary
    print("\n" + "=" * 80)
    print("CONVERSION SUMMARY")
    print("=" * 80)
    print(f"Total files processed:      {stats.total}")
    print(f"Successfully converted:     {stats.success}")
    print(f"Pre-validation failed:      {stats.pre_validation_failed}")
    print(f"Conversion failed:          {stats.conversion_failed}")
    print(f"Post-validation failed:     {stats.post_validation_failed}")
    print(f"Total failed:               {stats.total - stats.success}")
    print(f"Total warnings:             {len(stats.warnings)}")
    
    if stats.total > 0:
        success_rate = (stats.success / stats.total) * 100
        print(f"Success rate:               {success_rate:.2f}%")
    
    print("=" * 80)
    
    # Generate reports
    if not args.dry_run and (len(stats.errors) > 0 or len(stats.warnings) > 0):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Excel report
        excel_path = reports_dir / f"errors_{timestamp}.xlsx"
        generate_excel_report(stats, excel_path)
        
        # Text log
        log_path = logs_dir / f"conversion_{timestamp}.log"
        generate_text_log(stats, log_path)
    
    # Exit code
    if stats.success == stats.total:
        print("\n✓ All files converted successfully!")
        sys.exit(0)
    else:
        print(f"\n⚠ {stats.total - stats.success} files failed conversion.")
        print("Check the error reports for details.")
        sys.exit(1)


if __name__ == '__main__':
    main()