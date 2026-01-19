"""
Standalone validator for new JSON structure files.
Validates JSON files that are already in the new structure format.
Uses embedded validation logic from post_validator.py.
"""

import argparse
import sys
import json
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Tuple

try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False
    print("Warning: tqdm not installed. Progress bar will not be available.")
    print("Install with: pip install tqdm")

try:
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not installed. Excel reports will not be available.")
    print("Install with: pip install openpyxl")


# ============================================================================
# Configuration (from SCRIPTS/config.py)
# ============================================================================

LANGUAGES = {
    "en": "English",
    "ar": "Arabic",
    "de": "German",
    "fr": "French",
    "es": "Spanish",
    "it": "Italian",
    "pt": "Portuguese",
    "zh": "Chinese"
}

COUNTRIES = {
    "eg": "Egypt",
    "zz": "ZZ"
}


# ============================================================================
# Utility Functions (from SCRIPTS/utils.py)
# ============================================================================

class ValidationError(Exception):
    """Custom exception for validation errors"""
    def __init__(self, message: str, field: str = "", actual_value: Any = None, expected: str = ""):
        self.message = message
        self.field = field
        self.actual_value = actual_value
        self.expected = expected
        super().__init__(self.message)


def format_timestamp() -> str:
    """Return current timestamp in 'YYYY-MM-DD HH:MM:SS' format"""
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def load_json_file(filepath: Path) -> Dict[str, Any]:
    """
    Load JSON file with error handling.
    Raises ValidationError on failure.
    """
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data
    except json.JSONDecodeError as e:
        raise ValidationError(f"Invalid JSON: {str(e)}", "json", str(e), "Valid JSON")
    except Exception as e:
        raise ValidationError(f"Failed to read file: {str(e)}", "file", str(e), "Readable file")


# ============================================================================
# Validation Functions (from SCRIPTS/post_validator.py)
# ============================================================================

def _validate_root_fields(data: Dict[str, Any]) -> List[str]:
    """Validate root level fields of converted JSON"""
    errors = []
    
    # Required fields
    required_fields = [
        'question_id', 'language_code', 'language',
        'country_code', 'country', 'subject', 'grade',
        'number_of_parts', 'section_id', 'source', 'content'
    ]
    
    for field in required_fields:
        if field not in data:
            errors.append(f"Missing required root field: '{field}'")
    
    # Validate question_id is string
    if 'question_id' in data and not isinstance(data['question_id'], str):
        errors.append("'question_id' must be a string")
    
    # Validate parent_id (can be None or string)
    if 'parent_id' in data and data['parent_id'] is not None:
        if not isinstance(data['parent_id'], str):
            errors.append("'parent_id' must be null or a string")
    
    # Validate language_code
    if 'language_code' in data and data['language_code'] not in LANGUAGES:
        errors.append(f"Invalid language_code: '{data['language_code']}'")
    
    # Validate language matches language_code
    if 'language_code' in data and 'language' in data:
        expected_language = LANGUAGES.get(data['language_code'])
        if data['language'] != expected_language:
            errors.append(f"Language mismatch: got '{data['language']}', expected '{expected_language}'")
    
    # Validate country_code
    if 'country_code' in data and data['country_code'] is not None:
        if data['country_code'] not in COUNTRIES:
            errors.append(f"Invalid country_code: '{data['country_code']}'")
    
    # Validate country matches country_code
    if 'country_code' in data and 'country' in data:
        if data['country_code'] is not None:
            expected_country = COUNTRIES.get(data['country_code'])
            if data['country'] != expected_country:
                errors.append(f"Country mismatch: got '{data['country']}', expected '{expected_country}'")
    
    # Validate number_of_parts (must be integer, not string)
    if 'number_of_parts' in data:
        if not isinstance(data['number_of_parts'], int):
            errors.append(f"'number_of_parts' must be an integer, got {type(data['number_of_parts']).__name__}")
        elif data['number_of_parts'] < 1:
            errors.append("'number_of_parts' must be a positive integer")
    
    # Validate subject is string
    if 'subject' in data and not isinstance(data['subject'], str):
        errors.append(f"'subject' must be a string, got {type(data['subject']).__name__}")
    
    # Validate subject_id is string (required)
    if 'subject_id' not in data:
        errors.append("Missing required root field: 'subject_id'")
    elif not isinstance(data['subject_id'], str):
        errors.append(f"'subject_id' must be a string, got {type(data['subject_id']).__name__}")
    
    # Validate grade is string
    if 'grade' in data and not isinstance(data['grade'], str):
        errors.append(f"'grade' must be a string, got {type(data['grade']).__name__}")
    
    # Validate grade_id is string (required)
    if 'grade_id' not in data:
        errors.append("Missing required root field: 'grade_id'")
    elif not isinstance(data['grade_id'], str):
        errors.append(f"'grade_id' must be a string, got {type(data['grade_id']).__name__}")
    
    # Validate section_id is string
    if 'section_id' in data and not isinstance(data['section_id'], str):
        errors.append(f"'section_id' must be a string, got {type(data['section_id']).__name__}")
    
    # Validate source is string
    if 'source' in data and not isinstance(data['source'], str):
        errors.append(f"'source' must be a string, got {type(data['source']).__name__}")
    
    return errors


def _validate_content(content: Dict[str, Any], number_of_parts: Any) -> List[str]:
    """Validate content object"""
    errors = []
    
    # Validate parts array
    if 'parts' not in content:
        errors.append("Content missing 'parts' array")
        return errors
    
    if not isinstance(content['parts'], list):
        errors.append("'parts' must be an array")
        return errors
    
    if len(content['parts']) == 0:
        errors.append("'parts' array cannot be empty")
        return errors
    
    # Convert number_of_parts to int for comparison (handle string/type errors)
    try:
        num_parts_int = int(number_of_parts) if number_of_parts is not None else 0
    except (ValueError, TypeError):
        errors.append(f"'number_of_parts' must be an integer (got: {type(number_of_parts).__name__})")
        num_parts_int = 0  # Use 0 as fallback to continue validation
    
    # Validate statement logic
    if num_parts_int > 1:
        if 'statement' not in content:
            errors.append("Multi-part questions must have 'statement' in content")
    else:
        if 'statement' in content:
            errors.append("Single-part questions should not have 'statement' in content")
    
    # Validate number of parts matches
    if len(content['parts']) != num_parts_int:
        errors.append(f"Parts count mismatch: content has {len(content['parts'])} parts but number_of_parts is {num_parts_int}")
    
    # Validate each part
    for i, part in enumerate(content['parts'], 1):
        errors.extend(_validate_part(part, i, num_parts_int))
    
    return errors


def _validate_part(part: Dict[str, Any], part_number: int, total_parts: int) -> List[str]:
    """Validate a converted part"""
    errors = []
    
    # Required fields for all parts
    required_fields = ['n', 'type', 'stem']
    for field in required_fields:
        if field not in part:
            errors.append(f"Part {part_number}: Missing required field '{field}'")
    
    # Validate n is integer and matches position
    if 'n' in part:
        if not isinstance(part['n'], int):
            errors.append(f"Part {part_number}: 'n' must be an integer, got {type(part['n']).__name__}")
        elif part['n'] != part_number:
            errors.append(f"Part {part_number}: Part number 'n' ({part['n']}) does not match position")
    
    # Validate type is string
    if 'type' in part and not isinstance(part['type'], str):
        errors.append(f"Part {part_number}: 'type' must be a string, got {type(part['type']).__name__}")
    
    # Validate stem is string
    if 'stem' in part and not isinstance(part['stem'], str):
        errors.append(f"Part {part_number}: 'stem' must be a string, got {type(part['stem']).__name__}")
    
    # Validate explanation (if present) is string or null
    if 'explanation' in part and part['explanation'] is not None and not isinstance(part['explanation'], str):
        errors.append(f"Part {part_number}: 'explanation' must be a string or null, got {type(part['explanation']).__name__}")
    
    # Validate type-specific fields
    part_type = part.get('type')
    
    if part_type == 'counting':
        errors.extend(_validate_counting_part(part, part_number))
    elif part_type == 'frq':
        errors.extend(_validate_frq_part(part, part_number))
    elif part_type == 'gap':
        errors.extend(_validate_gap_part(part, part_number))
    elif part_type == 'input':
        errors.extend(_validate_input_part(part, part_number))
    elif part_type == 'matching':
        errors.extend(_validate_matching_part(part, part_number))
    elif part_type == 'gmrq':
        errors.extend(_validate_gmrq_part(part, part_number))
    elif part_type == 'mcq':
        errors.extend(_validate_mcq_part(part, part_number))
    elif part_type == 'mrq':
        errors.extend(_validate_mrq_part(part, part_number))
    elif part_type == 'opinion':
        errors.extend(_validate_opinion_part(part, part_number))
    elif part_type == 'ordering':
        errors.extend(_validate_ordering_part(part, part_number))
    elif part_type == 'puzzle':
        errors.extend(_validate_puzzle_part(part, part_number))
    elif part_type == 'string':
        errors.extend(_validate_string_part(part, part_number))
    
    return errors


def _validate_counting_part(part: Dict[str, Any], part_number: int) -> List[str]:
    """Validate counting type part"""
    errors = []
    
    if 'grid' not in part:
        errors.append(f"Part {part_number} (counting): Missing 'grid' object")
    elif isinstance(part['grid'], dict):
        if 'rows' not in part['grid'] or 'columns' not in part['grid']:
            errors.append(f"Part {part_number} (counting): 'grid' must have 'rows' and 'columns'")
        if not isinstance(part['grid'].get('rows'), int):
            errors.append(f"Part {part_number} (counting): 'grid.rows' must be an integer")
        if not isinstance(part['grid'].get('columns'), int):
            errors.append(f"Part {part_number} (counting): 'grid.columns' must be an integer")
    
    if 'correct_answer' not in part:
        errors.append(f"Part {part_number} (counting): Missing 'correct_answer'")
    elif not isinstance(part['correct_answer'], int):
        errors.append(f"Part {part_number} (counting): 'correct_answer' must be an integer")
    
    return errors


def _validate_frq_part(part: Dict[str, Any], part_number: int) -> List[str]:
    """Validate frq type part"""
    errors = []
    
    if 'acceptable_answers' not in part:
        errors.append(f"Part {part_number} (frq): Missing 'acceptable_answers'")
    elif not isinstance(part['acceptable_answers'], list):
        errors.append(f"Part {part_number} (frq): 'acceptable_answers' must be an array")

    # Validate ai_template_id
    if 'ai_template_id' not in part:
        errors.append(f"Part {part_number} (frq): 'ai_template_id' is required")
    else:
        t_id = part['ai_template_id']
        # 1. Check if it is a string
        if not isinstance(t_id, str):
            errors.append(f"Part {part_number} (frq): 'ai_template_id' must be a string")
        # 2. Check if it is exactly 12 digits
        elif not (t_id.isdigit() and len(t_id) == 12):
            errors.append(f"Part {part_number} (frq): 'ai_template_id' must be exactly 12 digits")

    return errors


def _validate_gap_part(part: Dict[str, Any], part_number: int) -> List[str]:
    """Validate gap type part"""
    errors = []
    
    if 'gap_keys' not in part:
        errors.append(f"Part {part_number} (gap): Missing 'gap_keys'")
    elif not isinstance(part['gap_keys'], list):
        errors.append(f"Part {part_number} (gap): 'gap_keys' must be an array")
    else:
        for i, key in enumerate(part['gap_keys']):
            if 'value' not in key:
                errors.append(f"Part {part_number} (gap): gap_key {i} missing 'value'")
            elif not isinstance(key.get('value'), str):
                errors.append(f"Part {part_number} (gap): gap_key {i} 'value' must be a string, got {type(key.get('value')).__name__}")
            if 'display_order' not in key:
                errors.append(f"Part {part_number} (gap): gap_key {i} missing 'display_order'")
            elif not isinstance(key.get('display_order'), int):
                errors.append(f"Part {part_number} (gap): gap_key {i} 'display_order' must be an integer, got {type(key.get('display_order')).__name__}")
            if 'correct_order' in key and not isinstance(key.get('correct_order'), int):
                errors.append(f"Part {part_number} (gap): gap_key {i} 'correct_order' must be an integer, got {type(key.get('correct_order')).__name__}")
    
    if 'correct_answer' not in part:
        errors.append(f"Part {part_number} (gap): Missing 'correct_answer'")
    elif not isinstance(part['correct_answer'], str):
        errors.append(f"Part {part_number} (gap): 'correct_answer' must be a string")
    
    return errors


def _validate_input_part(part: Dict[str, Any], part_number: int) -> List[str]:
    """Validate input type part"""
    errors = []
    
    if 'correct_answer' not in part:
        errors.append(f"Part {part_number} (input): Missing 'correct_answer'")
    elif not isinstance(part['correct_answer'], dict):
        errors.append(f"Part {part_number} (input): 'correct_answer' must be an object, got {type(part['correct_answer']).__name__}")
    else:
        if 'value' not in part['correct_answer']:
            errors.append(f"Part {part_number} (input): 'correct_answer.value' is required")
        elif not isinstance(part['correct_answer'].get('value'), (int, float)):
            errors.append(f"Part {part_number} (input): 'correct_answer.value' must be a number, got {type(part['correct_answer'].get('value')).__name__}")
        
        if 'unit' in part['correct_answer'] and part['correct_answer']['unit'] is not None:
            if not isinstance(part['correct_answer']['unit'], str):
                errors.append(f"Part {part_number} (input): 'correct_answer.unit' must be a string or null, got {type(part['correct_answer']['unit']).__name__}")
        
        if 'constraints' not in part['correct_answer']:
            errors.append(f"Part {part_number} (input): 'correct_answer.constraints' is required")
        elif not isinstance(part['correct_answer']['constraints'], dict):
            errors.append(f"Part {part_number} (input): 'correct_answer.constraints' must be an object, got {type(part['correct_answer']['constraints']).__name__}")
        else:
            constraints = part['correct_answer']['constraints']
            if 'type' not in constraints:
                errors.append(f"Part {part_number} (input): 'constraints.type' is required")
            elif not isinstance(constraints.get('type'), str):
                errors.append(f"Part {part_number} (input): 'constraints.type' must be a string, got {type(constraints.get('type')).__name__}")
            
            if 'answer_format' in constraints:
                if not isinstance(constraints['answer_format'], dict):
                    errors.append(f"Part {part_number} (input): 'constraints.answer_format' must be an object, got {type(constraints['answer_format']).__name__}")
                else:
                    af = constraints['answer_format']
                    if 'number_format' in af and not isinstance(af.get('number_format'), str):
                        errors.append(f"Part {part_number} (input): 'constraints.answer_format.number_format' must be a string, got {type(af.get('number_format')).__name__}")
                    if 'contains_thousand_separator' in af and not isinstance(af.get('contains_thousand_separator'), bool):
                        errors.append(f"Part {part_number} (input): 'constraints.answer_format.contains_thousand_separator' must be a boolean, got {type(af.get('contains_thousand_separator')).__name__}")
            
            # Validate boolean constraint fields
            bool_fields = [
                'allow_leading_zeros', 'allow_trailing_zeros',
                'remove_leading_zeros_from_normalized_form', 'remove_trailing_zeros_from_normalized_form',
                'add_single_leading_zero_to_normalized_form', 'remove_trailing_decimal_point_from_normalized_form'
            ]
            for field in bool_fields:
                if field in constraints and not isinstance(constraints.get(field), bool):
                    errors.append(f"Part {part_number} (input): 'constraints.{field}' must be a boolean, got {type(constraints.get(field)).__name__}")
            
            # Validate integer constraint fields (can be integer or null)
            int_fields = [
                'must_have_no_more_than_nsf', 'must_have_at_least_nsf', 'must_have_exactly_nsf',
                'must_have_at_least_ndp', 'must_have_no_more_than_ndp', 'must_have_exactly_ndp'
            ]
            for field in int_fields:
                if field in constraints and constraints[field] is not None:
                    if not isinstance(constraints[field], int):
                        errors.append(f"Part {part_number} (input): 'constraints.{field}' must be an integer or null, got {type(constraints[field]).__name__}")
            
            # Validate string constraint fields
            if 'sign' in constraints and not isinstance(constraints.get('sign'), str):
                errors.append(f"Part {part_number} (input): 'constraints.sign' must be a string, got {type(constraints.get('sign')).__name__}")
            if 'normalize_sign' in constraints and not isinstance(constraints.get('normalize_sign'), str):
                errors.append(f"Part {part_number} (input): 'constraints.normalize_sign' must be a string, got {type(constraints.get('normalize_sign')).__name__}")
    
    return errors


def _validate_matching_part(part: Dict[str, Any], part_number: int) -> List[str]:
    """Validate matching type part"""
    errors = []

    if 'correct_answer' not in part:
        errors.append(f"Part {part_number} (matching): Missing 'correct_answer'")        

    if 'items' not in part:
        errors.append(f"Part {part_number} (matching): Missing 'items'")
    elif isinstance(part['items'], dict):
        if 'A' not in part['items']:
            errors.append(f"Part {part_number} (matching): 'items.A' is required")
        if 'B' not in part['items']:
            errors.append(f"Part {part_number} (matching): 'items.B' is required")
    
    return errors


def _validate_gmrq_part(part: Dict[str, Any], part_number: int) -> List[str]:
    """Validate gmrq type part"""
    errors = []

    if 'correct_answer' not in part:
        errors.append(f"Part {part_number} (gmrq): Missing 'correct_answer'")        

    if 'items' not in part:
        errors.append(f"Part {part_number} (gmrq): Missing 'items'")
    elif isinstance(part['items'], dict):
        if 'A' not in part['items']:
            errors.append(f"Part {part_number} (gmrq): 'items.A' is required")
        if 'B' not in part['items']:
            errors.append(f"Part {part_number} (gmrq): 'items.B' is required")
    
    return errors


def _validate_mcq_part(part: Dict[str, Any], part_number: int) -> List[str]:
    """Validate mcq type part"""
    errors = []
    
    if 'choices' not in part:
        errors.append(f"Part {part_number} (mcq): Missing 'choices'")
    elif not isinstance(part['choices'], list):
        errors.append(f"Part {part_number} (mcq): 'choices' must be an array, got {type(part['choices']).__name__}")
    else:
        for i, choice in enumerate(part['choices']):
            if not isinstance(choice, dict):
                errors.append(f"Part {part_number} (mcq): choices[{i}] must be an object")
                continue
            if 'label' not in choice:
                errors.append(f"Part {part_number} (mcq): choices[{i}] missing 'label'")
            elif not isinstance(choice.get('label'), str):
                errors.append(f"Part {part_number} (mcq): choices[{i}] 'label' must be a string, got {type(choice.get('label')).__name__}")
            if 'value' not in choice:
                errors.append(f"Part {part_number} (mcq): choices[{i}] missing 'value'")
            elif not isinstance(choice.get('value'), str):
                errors.append(f"Part {part_number} (mcq): choices[{i}] 'value' must be a string, got {type(choice.get('value')).__name__}")
            if 'is_correct' not in choice:
                errors.append(f"Part {part_number} (mcq): choices[{i}] missing 'is_correct'")
            elif not isinstance(choice.get('is_correct'), bool):
                errors.append(f"Part {part_number} (mcq): choices[{i}] 'is_correct' must be a boolean, got {type(choice.get('is_correct')).__name__}")
    
    if 'correct_answer' not in part:
        errors.append(f"Part {part_number} (mcq): Missing 'correct_answer'")
    elif not isinstance(part['correct_answer'], dict):
        errors.append(f"Part {part_number} (mcq): 'correct_answer' must be an object, got {type(part['correct_answer']).__name__}")
    else:
        if 'label' not in part['correct_answer']:
            errors.append(f"Part {part_number} (mcq): 'correct_answer.label' is required")
        elif not isinstance(part['correct_answer'].get('label'), str):
            errors.append(f"Part {part_number} (mcq): 'correct_answer.label' must be a string, got {type(part['correct_answer'].get('label')).__name__}")
        if 'value' not in part['correct_answer']:
            errors.append(f"Part {part_number} (mcq): 'correct_answer.value' is required")
        elif not isinstance(part['correct_answer'].get('value'), str):
            errors.append(f"Part {part_number} (mcq): 'correct_answer.value' must be a string, got {type(part['correct_answer'].get('value')).__name__}")
    
    return errors


def _validate_mrq_part(part: Dict[str, Any], part_number: int) -> List[str]:
    """Validate mrq type part"""
    errors = []
    
    if 'choices' not in part:
        errors.append(f"Part {part_number} (mrq): Missing 'choices'")
    elif not isinstance(part['choices'], list):
        errors.append(f"Part {part_number} (mrq): 'choices' must be an array")
    
    if 'correct_answer' not in part:
        errors.append(f"Part {part_number} (mrq): Missing 'correct_answer'")
    elif not isinstance(part['correct_answer'], list):
        errors.append(f"Part {part_number} (mrq): 'correct_answer' must be an array")
    
    return errors


def _validate_opinion_part(part: Dict[str, Any], part_number: int) -> List[str]:
    """Validate opinion type part"""
    errors = []
    
    if 'choices' not in part:
        errors.append(f"Part {part_number} (opinion): Missing 'choices'")
    elif not isinstance(part['choices'], list):
        errors.append(f"Part {part_number} (opinion): 'choices' must be an array")
    
    # Opinion should NOT have correct_answer
    if 'correct_answer' in part:
        errors.append(f"Part {part_number} (opinion): Should not have 'correct_answer'")
    
    return errors


def _validate_ordering_part(part: Dict[str, Any], part_number: int) -> List[str]:
    """Validate ordering type part"""
    errors = []
    
    if 'direction' not in part:
        errors.append(f"Part {part_number} (ordering): Missing 'direction'")
    elif not isinstance(part['direction'], str):
        errors.append(f"Part {part_number} (ordering): 'direction' must be a string, got {type(part['direction']).__name__}")
    elif part['direction'] not in ['vertical', 'horizontal']:
        errors.append(f"Part {part_number} (ordering): 'direction' must be 'vertical' or 'horizontal'")
    
    if 'items' not in part:
        errors.append(f"Part {part_number} (ordering): Missing 'items'")
    elif not isinstance(part['items'], list):
        errors.append(f"Part {part_number} (ordering): 'items' must be an array, got {type(part['items']).__name__}")
    else:
        for i, item in enumerate(part['items']):
            if not isinstance(item, dict):
                errors.append(f"Part {part_number} (ordering): items[{i}] must be an object")
                continue
            if 'value' not in item:
                errors.append(f"Part {part_number} (ordering): items[{i}] missing 'value'")
            elif not isinstance(item.get('value'), str):
                errors.append(f"Part {part_number} (ordering): items[{i}] 'value' must be a string, got {type(item.get('value')).__name__}")
            if 'display_order' not in item:
                errors.append(f"Part {part_number} (ordering): items[{i}] missing 'display_order'")
            elif not isinstance(item.get('display_order'), int):
                errors.append(f"Part {part_number} (ordering): items[{i}] 'display_order' must be an integer, got {type(item.get('display_order')).__name__}")
            if 'correct_order' not in item:
                errors.append(f"Part {part_number} (ordering): items[{i}] missing 'correct_order'")
            elif not isinstance(item.get('correct_order'), int):
                errors.append(f"Part {part_number} (ordering): items[{i}] 'correct_order' must be an integer, got {type(item.get('correct_order')).__name__}")
    
    if 'correct_answer' not in part:
        errors.append(f"Part {part_number} (ordering): Missing 'correct_answer'")
    elif not isinstance(part['correct_answer'], list):
        errors.append(f"Part {part_number} (ordering): 'correct_answer' must be an array, got {type(part['correct_answer']).__name__}")
    
    return errors


def _validate_puzzle_part(part: Dict[str, Any], part_number: int) -> List[str]:
    """Validate puzzle type part"""
    errors = []
    
    if 'rows' not in part:
        errors.append(f"Part {part_number} (puzzle): Missing 'rows'")
    elif not isinstance(part['rows'], int):
        errors.append(f"Part {part_number} (puzzle): 'rows' must be an integer, got {type(part['rows']).__name__}")
    
    if 'columns' not in part:
        errors.append(f"Part {part_number} (puzzle): Missing 'columns'")
    elif not isinstance(part['columns'], int):
        errors.append(f"Part {part_number} (puzzle): 'columns' must be an integer, got {type(part['columns']).__name__}")
    
    if 'pieces' not in part:
        errors.append(f"Part {part_number} (puzzle): Missing 'pieces'")
    elif not isinstance(part['pieces'], list):
        errors.append(f"Part {part_number} (puzzle): 'pieces' must be an array, got {type(part['pieces']).__name__}")
    else:
        for i, piece in enumerate(part['pieces']):
            if not isinstance(piece, dict):
                errors.append(f"Part {part_number} (puzzle): pieces[{i}] must be an object")
                continue
            if 'display_order' not in piece:
                errors.append(f"Part {part_number} (puzzle): pieces[{i}] missing 'display_order'")
            elif not isinstance(piece.get('display_order'), int):
                errors.append(f"Part {part_number} (puzzle): pieces[{i}] 'display_order' must be an integer, got {type(piece.get('display_order')).__name__}")
            if 'correct_order' not in piece:
                errors.append(f"Part {part_number} (puzzle): pieces[{i}] missing 'correct_order'")
            elif not isinstance(piece.get('correct_order'), int):
                errors.append(f"Part {part_number} (puzzle): pieces[{i}] 'correct_order' must be an integer, got {type(piece.get('correct_order')).__name__}")
            if 'src' not in piece:
                errors.append(f"Part {part_number} (puzzle): pieces[{i}] missing 'src'")
            elif not isinstance(piece.get('src'), str):
                errors.append(f"Part {part_number} (puzzle): pieces[{i}] 'src' must be a string, got {type(piece.get('src')).__name__}")
            if 'alt' not in piece:
                errors.append(f"Part {part_number} (puzzle): pieces[{i}] missing 'alt'")
            elif not isinstance(piece.get('alt'), str):
                errors.append(f"Part {part_number} (puzzle): pieces[{i}] 'alt' must be a string, got {type(piece.get('alt')).__name__}")
    
    if 'correct_answer' not in part:
        errors.append(f"Part {part_number} (puzzle): Missing 'correct_answer'")
    elif not isinstance(part['correct_answer'], dict):
        errors.append(f"Part {part_number} (puzzle): 'correct_answer' must be an object, got {type(part['correct_answer']).__name__}")
    else:
        if 'src' not in part['correct_answer']:
            errors.append(f"Part {part_number} (puzzle): 'correct_answer.src' is required")
        elif not isinstance(part['correct_answer'].get('src'), str):
            errors.append(f"Part {part_number} (puzzle): 'correct_answer.src' must be a string, got {type(part['correct_answer'].get('src')).__name__}")
        if 'alt' not in part['correct_answer']:
            errors.append(f"Part {part_number} (puzzle): 'correct_answer.alt' is required")
        elif not isinstance(part['correct_answer'].get('alt'), str):
            errors.append(f"Part {part_number} (puzzle): 'correct_answer.alt' must be a string, got {type(part['correct_answer'].get('alt')).__name__}")
    
    return errors


def _validate_string_part(part: Dict[str, Any], part_number: int) -> List[str]:
    """Validate string type part"""
    errors = []
    
    if 'ai_template_id' not in part:
        errors.append(f"Part {part_number} (string): Missing 'ai_template_id'")
    elif not isinstance(part['ai_template_id'], str):
        errors.append(f"Part {part_number} (string): 'ai_template_id' must be a string, got {type(part['ai_template_id']).__name__}")
    
    if 'acceptable_answers' not in part:
        errors.append(f"Part {part_number} (string): Missing 'acceptable_answers'")
    elif not isinstance(part['acceptable_answers'], list):
        errors.append(f"Part {part_number} (string): 'acceptable_answers' must be an array, got {type(part['acceptable_answers']).__name__}")
    else:
        for i, answer in enumerate(part['acceptable_answers']):
            if not isinstance(answer, str):
                errors.append(f"Part {part_number} (string): acceptable_answers[{i}] must be a string, got {type(answer).__name__}")
    
    if 'guidelines' in part:
        if not isinstance(part['guidelines'], list):
            errors.append(f"Part {part_number} (string): 'guidelines' must be an array, got {type(part['guidelines']).__name__}")
        else:
            for i, guideline in enumerate(part['guidelines']):
                if not isinstance(guideline, dict):
                    errors.append(f"Part {part_number} (string): guidelines[{i}] must be an object")
                    continue
                if 'student_answer' not in guideline:
                    errors.append(f"Part {part_number} (string): guidelines[{i}] missing 'student_answer'")
                elif not isinstance(guideline.get('student_answer'), str):
                    errors.append(f"Part {part_number} (string): guidelines[{i}] 'student_answer' must be a string, got {type(guideline.get('student_answer')).__name__}")
                if 'mark' not in guideline:
                    errors.append(f"Part {part_number} (string): guidelines[{i}] missing 'mark'")
                elif not isinstance(guideline.get('mark'), (int, float)):
                    errors.append(f"Part {part_number} (string): guidelines[{i}] 'mark' must be a number, got {type(guideline.get('mark')).__name__}")
                elif guideline.get('mark') not in [0, 1, 0.0, 1.0]:
                    errors.append(f"Part {part_number} (string): guidelines[{i}] 'mark' must be 0 or 1")
                if 'comment' not in guideline:
                    errors.append(f"Part {part_number} (string): guidelines[{i}] missing 'comment'")
                elif not isinstance(guideline.get('comment'), str):
                    errors.append(f"Part {part_number} (string): guidelines[{i}] 'comment' must be a string, got {type(guideline.get('comment')).__name__}")
    
    return errors


def validate_post_conversion(converted_json: Dict[str, Any]) -> Tuple[bool, List[str]]:
    """
    Main entry point for post-conversion validation.
    Validates that the converted JSON has correct structure and all required fields.
    Returns: Tuple of (is_valid, list_of_errors)
    """
    errors = []
    
    # Validate root level fields
    errors.extend(_validate_root_fields(converted_json))
    
    # Validate content structure
    if 'content' in converted_json:
        number_of_parts = converted_json.get('number_of_parts', 0)
        errors.extend(_validate_content(converted_json['content'], number_of_parts))
    else:
        errors.append("Missing 'content' object")
    
    return (len(errors) == 0, errors)


# ============================================================================
# Main Validation Script
# ============================================================================

class ValidationStats:
    """Track validation statistics"""
    def __init__(self):
        self.total = 0
        self.valid = 0
        self.invalid = 0
        self.errors = []  # List of error dictionaries for Excel report
        
    def add_error(self, filename: str, question_id: str, error_message: str):
        """Add an error to the error log"""
        self.errors.append({
            "filename": filename,
            "question_id": question_id,
            "error_message": error_message,
            "timestamp": format_timestamp()
        })


def discover_json_files(input_path: Path) -> List[Path]:
    """
    Discover all JSON files in input path recursively.
    
    Args:
        input_path: Root directory to search or single file path
        
    Returns:
        List of JSON file paths
    """
    json_files = []
    
    if input_path.is_file():
        if input_path.suffix.lower() == '.json':
            json_files.append(input_path)
    elif input_path.is_dir():
        json_files = list(input_path.rglob('*.json'))
    else:
        raise ValueError(f"Input path does not exist: {input_path}")
    
    return sorted(json_files)


def validate_file(filepath: Path, stats: ValidationStats, verbose: bool = False) -> bool:
    """
    Validate a single JSON file using post_validator.
    
    Args:
        filepath: Path to the JSON file
        stats: Statistics tracker
        verbose: If True, print detailed messages
        
    Returns:
        True if valid, False otherwise
    """
    filename = filepath.name
    
    try:
        # Load JSON file
        json_data = load_json_file(filepath)
        question_id = str(json_data.get('question_id', 'unknown'))
        
        # Validate using post_validator
        is_valid, errors = validate_post_conversion(json_data)
        
        if is_valid:
            stats.valid += 1
            if verbose:
                print(f"  ✓ VALID: {filename}")
            return True
        else:
            stats.invalid += 1
            
            # Log all errors
            for error in errors:
                stats.add_error(filename, question_id, error)
            
            if verbose:
                print(f"  ✗ INVALID: {filename} - {len(errors)} error(s)")
                for error in errors[:3]:  # Show first 3 errors
                    print(f"    - {error}")
                if len(errors) > 3:
                    print(f"    ... and {len(errors) - 3} more error(s)")
            
            return False
            
    except ValidationError as e:
        stats.invalid += 1
        question_id = 'unknown'
        try:
            json_data = load_json_file(filepath)
            question_id = str(json_data.get('question_id', 'unknown'))
        except:
            pass
        
        error_msg = f"Failed to load JSON file: {str(e)}"
        stats.add_error(filename, question_id, error_msg)
        
        if verbose:
            print(f"  ✗ ERROR: {filename} - {error_msg}")
        
        return False
        
    except Exception as e:
        stats.invalid += 1
        question_id = 'unknown'
        try:
            json_data = load_json_file(filepath)
            question_id = str(json_data.get('question_id', 'unknown'))
        except:
            pass
        
        error_msg = f"Unexpected error: {str(e)}"
        stats.add_error(filename, question_id, error_msg)
        
        if verbose:
            print(f"  ✗ ERROR: {filename} - {error_msg}")
        
        return False


def generate_excel_report(stats: ValidationStats, output_path: Path):
    """
    Generate Excel report with validation errors.
    
    Args:
        stats: Statistics object
        output_path: Path to save the Excel file
    """
    if not HAS_OPENPYXL:
        print("Warning: openpyxl not available. Skipping Excel report generation.")
        return
    
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create Errors sheet
    ws_errors = wb.create_sheet("Errors")
    
    # Write headers
    headers = ["Filename", "Question ID", "Error Message", "Timestamp"]
    for col_idx, header in enumerate(headers, 1):
        ws_errors.cell(row=1, column=col_idx, value=header)
        # Bold headers
        ws_errors.cell(row=1, column=col_idx).font = ws_errors.cell(row=1, column=col_idx).font.copy(bold=True)
    
    # Write errors
    for row_idx, error in enumerate(stats.errors, 2):
        ws_errors.cell(row=row_idx, column=1, value=error['filename'])
        ws_errors.cell(row=row_idx, column=2, value=error['question_id'])
        ws_errors.cell(row=row_idx, column=3, value=error['error_message'])
        ws_errors.cell(row=row_idx, column=4, value=error['timestamp'])
    
    # Auto-adjust column widths
    for column in ws_errors.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)
        ws_errors.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)
    print(f"\nExcel report saved to: {output_path}")
    print(f"  - {len(stats.errors)} errors in 'Errors' sheet")


def generate_text_log(stats: ValidationStats, output_path: Path):
    """
    Generate text log of validation process.
    
    Args:
        stats: Statistics object
        output_path: Path to save the log file
    """
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write("NEW STRUCTURE VALIDATION LOG\n")
        f.write("=" * 80 + "\n\n")
        f.write(f"Generated: {format_timestamp()}\n\n")
        
        f.write("SUMMARY\n")
        f.write("-" * 80 + "\n")
        f.write(f"Total files processed: {stats.total}\n")
        f.write(f"Valid files: {stats.valid}\n")
        f.write(f"Invalid files: {stats.invalid}\n")
        f.write(f"Total errors: {len(stats.errors)}\n\n")
        
        if len(stats.errors) > 0:
            f.write("ERRORS\n")
            f.write("-" * 80 + "\n")
            
            # Group errors by filename
            errors_by_file: Dict[str, List[Dict[str, Any]]] = {}
            for error in stats.errors:
                filename = error['filename']
                if filename not in errors_by_file:
                    errors_by_file[filename] = []
                errors_by_file[filename].append(error)
            
            for filename in sorted(errors_by_file.keys()):
                file_errors = errors_by_file[filename]
                question_id = file_errors[0]['question_id']
                f.write(f"\nFile: {filename}\n")
                f.write(f"Question ID: {question_id}\n")
                f.write(f"Errors ({len(file_errors)}):\n")
                for error in file_errors:
                    f.write(f"  - {error['error_message']}\n")
                f.write(f"Timestamp: {file_errors[0]['timestamp']}\n")
        
        f.write("\n" + "=" * 80 + "\n")
        f.write("End of validation log\n")
        f.write("=" * 80 + "\n")
    
    print(f"Text log saved to: {output_path}")


def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(
        description='Validate JSON files in new structure format',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Validate all JSON files in a directory
  python3 validate_new_structure.py --input path/to/json/files
  
  # Validate with verbose output
  python3 validate_new_structure.py --input path/to/json/files -v
  
  # Specify custom output directory for reports
  python3 validate_new_structure.py --input path/to/json/files --output validation_reports
        """
    )
    
    parser.add_argument(
        '--input', '-i',
        type=str,
        default='OUTPUTS/CONVERTED',
        help='Input directory or file path containing JSON files to validate (default: OUTPUTS/CONVERTED)'
    )
    
    parser.add_argument(
        '--output', '-o',
        type=str,
        default='VALIDATION_REPORTS',
        help='Output directory for validation reports (default: VALIDATION_REPORTS)'
    )
    
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Verbose mode: show detailed validation results for each file'
    )
    
    args = parser.parse_args()
    
    # Setup paths
    input_path = Path(args.input)
    output_dir = Path(args.output)
    
    if not input_path.exists():
        print(f"Error: Input path does not exist: {input_path}")
        sys.exit(1)
    
    # Create output directory
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Discover JSON files
    print(f"Discovering JSON files in: {input_path}")
    json_files = discover_json_files(input_path)
    
    if not json_files:
        print(f"No JSON files found in: {input_path}")
        sys.exit(0)
    
    print(f"Found {len(json_files)} JSON file(s) to validate\n")
    
    # Initialize statistics
    stats = ValidationStats()
    stats.total = len(json_files)
    
    # Validate files
    print("Validating files...")
    if HAS_TQDM:
        for filepath in tqdm(json_files, desc="Validating", unit="file"):
            validate_file(filepath, stats, args.verbose)
    else:
        for filepath in json_files:
            validate_file(filepath, stats, args.verbose)
    
    # Generate reports
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Excel report
    excel_path = output_dir / f"validation_report_{timestamp}.xlsx"
    generate_excel_report(stats, excel_path)
    
    # Text log
    log_path = output_dir / f"validation_log_{timestamp}.txt"
    generate_text_log(stats, log_path)
    
    # Print summary
    print("\n" + "=" * 80)
    print("VALIDATION SUMMARY")
    print("=" * 80)
    print(f"Total files processed: {stats.total}")
    print(f"Valid files: {stats.valid}")
    print(f"Invalid files: {stats.invalid}")
    print(f"Total errors: {len(stats.errors)}")
    
    if stats.invalid > 0:
        print(f"\nValidation completed with errors. Check reports in: {output_dir}")
        sys.exit(1)
    else:
        print("\nAll files are valid!")
        sys.exit(0)


if __name__ == '__main__':
    main()
